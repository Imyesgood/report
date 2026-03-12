"""
parser.py
엑셀 데이터 파싱 + yfinance GOLD 데이터 → data.json 생성
"""

import openpyxl
import json
from datetime import datetime, date, timedelta
from pathlib import Path


# ── 설정 로드 ──────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
with open(BASE_DIR / "config.json", encoding="utf-8") as f:
    CONFIG = json.load(f)

EXCEL_PATH = BASE_DIR / CONFIG["excel_path"]
MAX_ROWS   = CONFIG.get("max_rows", 500)


# ── 시트 파싱 ──────────────────────────────────────────────────────────────────
def parse_sheet(wb, sheet_name):
    """
    2행 헤더 구조 시트를 파싱.
    반환: {인덱스명: {"cols": [...], "rows": [...]}}
    - row1: 인덱스명 (단위: 로 시작하는 건 무시)
    - row2: 컬럼명
    - row3~: 데이터 (최신이 위)
    """
    if sheet_name not in wb.sheetnames:
        print(f"[경고] 시트 없음: {sheet_name}")
        return {}

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, MAX_ROWS + 2), values_only=True))
    if len(all_rows) < 3:
        return {}

    row1 = all_rows[0]
    row2 = all_rows[1]
    data_rows = all_rows[2:]

    # 인덱스 그룹 위치 찾기 (단위: 제외, 중복 첫 번째만)
    groups = []
    seen = set()
    for i, val in enumerate(row1):
        if val is not None and not str(val).startswith("단위:"):
            if val not in seen:
                groups.append((str(val).strip(), i))
                seen.add(val)

    result = {}
    for g_idx, (name, start_col) in enumerate(groups):
        # 다음 그룹 시작 전까지
        if g_idx + 1 < len(groups):
            end_col = groups[g_idx + 1][1]
        else:
            end_col = len(row1)

        # 컬럼명 추출
        cols = []
        for c in row2[start_col:end_col]:
            cols.append(str(c).strip() if c is not None else f"_col{len(cols)}")

        # 데이터 행 추출
        rows = []
        for row in data_rows:
            row_slice = row[start_col:end_col]
            if any(v is not None for v in row_slice):
                rows.append(list(row_slice))

        if rows and cols:
            result[name] = {"cols": cols, "rows": rows}

    return result


def sheet_to_series(sheet_data, index_name, value_col, fallback_col=None):
    """
    시트 데이터에서 특정 인덱스의 {date: value} 시계열 추출.
    fallback_col: value_col이 None이면 대신 사용할 컬럼
    """
    if index_name not in sheet_data:
        print(f"[경고] 인덱스 없음: {index_name}")
        return {}

    info = sheet_data[index_name]
    cols = info["cols"]
    rows = info["rows"]

    if "일자" not in cols:
        print(f"[경고] 일자 컬럼 없음: {index_name}")
        return {}

    date_idx  = cols.index("일자")
    value_idx = cols.index(value_col) if value_col in cols else None
    fb_idx    = cols.index(fallback_col) if (fallback_col and fallback_col in cols) else None

    if value_idx is None and fb_idx is None:
        print(f"[경고] 값 컬럼 없음: {index_name} / {value_col}")
        return {}

    series = {}
    for row in rows:
        raw_date = row[date_idx]
        if raw_date is None:
            continue

        # 날짜 파싱
        if isinstance(raw_date, (datetime,)):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            try:
                d = datetime.fromisoformat(str(raw_date)).date()
            except Exception:
                continue

        # 값 추출 (value_col 우선, None이면 fallback)
        val = row[value_idx] if value_idx is not None else None
        if val is None and fb_idx is not None:
            val = row[fb_idx]

        if val is not None:
            try:
                series[d] = float(val)
            except (TypeError, ValueError):
                pass

    return series


# ── 날짜 계산 ──────────────────────────────────────────────────────────────────
def find_prev(series, before_date):
    """before_date 이전 중 가장 최근 (date, value) 반환"""
    candidates = [(d, v) for d, v in series.items() if d < before_date]
    if not candidates:
        return None, None
    candidates.sort(reverse=True)
    return candidates[0]


def find_nearest_on_or_before(series, target_date):
    """target_date 이하 중 가장 가까운 (date, value) 반환"""
    candidates = [(d, v) for d, v in series.items() if d <= target_date]
    if not candidates:
        return None, None
    candidates.sort(reverse=True)
    return candidates[0]


def find_nearest_on_or_after(series, target_date):
    """target_date 이상 중 가장 가까운 (date, value) 반환"""
    candidates = [(d, v) for d, v in series.items() if d >= target_date]
    if not candidates:
        return None, None
    candidates.sort()
    return candidates[0]


def calculate_index_data(series, today):
    """
    T0/1D/1M/YTM 계산.
    오늘 오전 기준: T0 = today 이전 가장 최근 영업일
    """
    if not series:
        return make_holiday()

    # T0
    t0_date, t0_val = find_prev(series, today)
    if t0_val is None:
        return make_holiday()

    # 1D: T0 이전 영업일
    t1_date, t1_val = find_prev(series, t0_date)
    d1 = round(t0_val - t1_val, 4) if t1_val is not None else None

    # 1M: T0 기준 30일 전
    target_1m = t0_date - timedelta(days=30)
    _, m1_base = find_nearest_on_or_before(series, target_1m)
    m1 = round(t0_val - m1_base, 4) if m1_base is not None else None

    # YTM: 올해 1/1 또는 그 이후 첫 거래일
    jan1 = date(today.year, 1, 1)
    _, ytm_base = find_nearest_on_or_after(series, jan1)
    ytm = round(t0_val - ytm_base, 4) if ytm_base is not None else None

    return {
        "holiday": False,
        "t0_date": t0_date.isoformat(),
        "t0":  round(t0_val, 4),
        "d1":  d1,
        "m1":  m1,
        "ytm": ytm,
    }


def make_holiday():
    return {"holiday": True, "t0": None, "d1": None, "m1": None, "ytm": None}


# ── yfinance GOLD ──────────────────────────────────────────────────────────────
def fetch_gold(today):
    try:
        import yfinance as yf
        ticker = yf.Ticker("GC=F")
        hist = ticker.history(period="2y", auto_adjust=True)
        if hist.empty:
            return make_holiday()

        series = {}
        for idx, row in hist.iterrows():
            d = idx.date() if hasattr(idx, 'date') else idx
            series[d] = float(row['Close'])

        return calculate_index_data(series, today)

    except ImportError:
        print("[경고] yfinance 없음. pip install yfinance 필요")
        return {**make_holiday(), "error": "yfinance not installed"}
    except Exception as e:
        print(f"[경고] GOLD 데이터 오류: {e}")
        return {**make_holiday(), "error": str(e)}


# ── 메인 파서 ──────────────────────────────────────────────────────────────────
def parse_all(today=None):
    if today is None:
        today = date.today()

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"엑셀 파일 없음: {EXCEL_PATH}")

    print(f"엑셀 로딩: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    # 시트별 파싱 (중복 파싱 방지)
    sheet_cache = {}
    def get_sheet(name):
        if name not in sheet_cache:
            sheet_cache[name] = parse_sheet(wb, name)
        return sheet_cache[name]

    result = {
        "generated_at": datetime.now().isoformat(),
        "base_date": None,
        "indices_left": [],
        "indices_right": [],
    }

    all_t0_dates = []

    def process_index(idx_cfg):
        sheet = idx_cfg["sheet"]

        if sheet == "yfinance":
            data = fetch_gold(today)
        else:
            sheet_data = get_sheet(sheet)
            series = sheet_to_series(
                sheet_data,
                idx_cfg["index"],
                idx_cfg["value_col"],
                idx_cfg.get("fallback_col"),
            )
            data = calculate_index_data(series, today)

        if data.get("t0_date"):
            all_t0_dates.append(data["t0_date"])

        return {
            "name": idx_cfg["name"],
            "fmt":  idx_cfg.get("fmt", "price"),
            "unit": idx_cfg.get("unit", ""),
            **data,
        }

    print("좌측 지수 파싱...")
    for cfg in CONFIG["indices_left"]:
        entry = process_index(cfg)
        result["indices_left"].append(entry)
        status = "휴장" if entry["holiday"] else f"{entry['t0']}"
        print(f"  {cfg['name']:20s} {status}")

    print("우측 지수 파싱...")
    for cfg in CONFIG["indices_right"]:
        entry = process_index(cfg)
        result["indices_right"].append(entry)
        status = "휴장" if entry["holiday"] else f"{entry['t0']}"
        print(f"  {cfg['name']:20s} {status}")

    wb.close()

    # 기준일: 가장 많이 등장한 T0 날짜
    if all_t0_dates:
        from collections import Counter
        most_common = Counter(all_t0_dates).most_common(1)[0][0]
        result["base_date"] = most_common

    # data.json 저장
    out_path = BASE_DIR / "data" / "data.json"
    out_path.parent.mkdir(exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n저장 완료: {out_path}")
    print(f"기준일: {result['base_date']}")
    return result


if __name__ == "__main__":
    parse_all()

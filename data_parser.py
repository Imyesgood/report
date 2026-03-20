
"""
data_parser.py — 사용자 지정 날짜 절대기준 파서
- T0*, T0, T-1, 연초는 사용자가 지정한 값을 그대로 사용
- T0 / T-1 / 연초는 exact match만 허용
- 값이 없는 지표는 그 지표만 pending 처리
- 1M은 T0-1개월 exact match 시에만 표시 (없어도 pending 사유로 보지 않음)
"""

import json
import os
from datetime import date, datetime, timedelta

import openpyxl
from dateutil.relativedelta import relativedelta


INDEX_CONFIG = [
    {"label":"통안 2Y",      "section":"left",  "type":"rate",
     "sheet":"국내채권", "header":"통안 2Y",           "value_col":"민평3사 수익률(산출일) 당일"},
    {"label":"국고 3Y",      "section":"left",  "type":"rate",
     "sheet":"국내채권", "header":"국고 3Y",            "value_col":"민평3사 수익률(산출일) 당일"},
    {"label":"국고 5Y",      "section":"left",  "type":"rate",
     "sheet":"국내채권", "header":"국고 5Y",            "value_col":"민평3사 수익률(산출일) 당일"},
    {"label":"국고 10Y",     "section":"left",  "type":"rate",
     "sheet":"국내채권", "header":"국고 10Y",           "value_col":"민평3사 수익률(산출일) 당일"},
    {"label":"국채3년선물",  "section":"left",  "type":"equity",
     "sheet":"국내채권", "header":"3년국채 연결",        "value_col":"현재가"},

    {"label":"USD/KRW",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"서울외환(기업용) USDKRW 스팟 (~15:30)", "value_col":"현재가"},
    {"label":"NDF",          "section":"left",  "type":"fx",
     "sheet":"환율", "header":"NDF 뉴욕 NDF 뉴욕",     "value_col":"NDF_MID_Close"},
    {"label":"Dollar Index", "section":"left",  "type":"fx",
     "sheet":"환율", "header":"달러인덱스 DOLLARS",     "value_col":"KR_MID_Close"},
    {"label":"USD/JPY",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"서울외환 이종통화 USDJPY","value_col":"Close"},
    {"label":"EUR/USD",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"서울외환 이종통화 EURUSD","value_col":"Close"},
    {"label":"JPY/KRW",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"서울외환 이종통화 JPYKRW","value_col":"Close"},
    {"label":"USD/CNY",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"중국:USDCNY:뉴욕종가",   "value_col":"현재가"},
    {"label":"GBP/USD",      "section":"left",  "type":"fx",
     "sheet":"환율", "header":"영국:GBPUSD:뉴욕종가",   "value_col":"현재가"},

    {"label":"KOSPI",        "section":"right", "type":"equity",
     "sheet":"주가지수", "header":"KOSPI",              "value_col":"현재가"},
    {"label":"NIKKEI",       "section":"right", "type":"equity",
     "sheet":"주가지수", "header":"니케이 225",          "value_col":"현재가", "ytm_date":"2026-01-05"},
    {"label":"중국상해종합",  "section":"right", "type":"equity",
     "sheet":"지수",    "header":"중국:상하이종합지수",   "value_col":"현재가"},
    {"label":"DOW",          "section":"right", "type":"equity",
     "sheet":"주가지수", "header":"다우 종합",                "value_col":"현재가"},
    {"label":"S&P500",       "section":"right", "type":"equity",
     "sheet":"주가지수", "header":"S&P 500",             "value_col":"현재가"},
    {"label":"NASDAQ",       "section":"right", "type":"equity",
     "sheet":"주가지수", "header":"나스닥",              "value_col":"현재가"},

    {"label":"T-Note (2yr)", "section":"right", "type":"rate",
     "sheet":"해외채권", "header":"2년 T-NOTE",          "value_col":"현재가"},
    {"label":"T-Note (10yr)","section":"right", "type":"rate",
     "sheet":"해외채권", "header":"10년 T-NOTE",         "value_col":"현재가"},
    {"label":"T-Bond (30yr)","section":"right", "type":"rate",
     "sheet":"해외채권", "header":"30년 T-BOND",         "value_col":"현재가"},

    {"label":"Germany 10Y", "section":"left", "type":"rate",
     "sheet":"해외채권",
     "header_candidates":["10년 독일 BUND","독일 10년 분트","10년 BUND","독일:10년분트","독일국채10년","BUND 10Y"],
     "header":"10년 독일 BUND", "value_col":"현재가"},
    {"label":"UK 10Y",      "section":"left", "type":"rate",
     "sheet":"해외채권",
     "header_candidates":["10년 영국 GILT","영국 10년 길트","10년 GILT","영국:10년길트","영국국채10년","GILT 10Y"],
     "header":"10년 영국 GILT", "value_col":"현재가"},
    {"label":"Japan 10Y",   "section":"left", "type":"rate",
     "sheet":"해외채권",
     "header_candidates":["10년 일본 JGB","일본 10년 국채","10년 JGB","일본:10년국채","일본국채10년","JGB 10Y"],
     "header":"10년 일본 JGB", "value_col":"현재가", "ytm_date":"2026-01-05"},

    {"label":"WTI",          "section":"left",  "type":"commodity",
     "sheet":"원자재",  "header":"WTI 현물",             "value_col":"현재가"},
    {"label":"GOLD",         "section":"left",  "type":"commodity",
     "sheet":"원자재",   "header":"금 고시가격 USD 온스 AM",  "value_col":"현재가"},
    {"label":"SOFR",         "section":"left",  "type":"rate",
     "sheet":"외환",    "header":"미국:SOFR:90일평균",   "value_col":"현재가"},
]

CHART_CONFIG = [
    {"label": "Korea 10Y",   "color": "#14422e", "dash": "solid",
     "sheet": "국내채권", "header": "국고 10Y",    "value_col": "민평3사 수익률(산출일) 당일"},
    {"label": "US 10Y",      "color": "#1d4ed8", "dash": "dashed",
     "sheet": "해외채권", "header": "10년 T-NOTE", "value_col": "현재가"},
    {"label": "Germany 10Y", "color": "#b45309", "dash": "dotted",
     "sheet": "해외채권",
     "header_candidates": ["10년 독일 BUND","독일 10년 분트","10년 BUND","독일:10년분트","독일국채10년","BUND 10Y"],
     "header": "10년 독일 BUND", "value_col": "현재가"},
]


def parse_iso_date(value, field_name):
    if value in (None, ""):
        raise ValueError(f"{field_name} 값이 비어 있습니다.")
    try:
        return date.fromisoformat(str(value).strip())
    except Exception as exc:
        raise ValueError(f"{field_name} 날짜 형식이 올바르지 않습니다: {value}") from exc


def find_columns(ws, header_text, value_col_text):
    header_col = None
    target_header = str(header_text).strip()
    target_value = str(value_col_text).strip()
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == target_header:
            header_col = cell.column
            break
    if header_col is None:
        return None, None

    date_col = None
    for offset in [0, 1, -1, 2, -2]:
        col = header_col + offset
        if col < 1:
            continue
        if ws.cell(row=2, column=col).value == "일자":
            date_col = col
            break
    if date_col is None:
        return None, None

    val_col = None
    for col in range(date_col + 1, date_col + 16):
        cell_val = ws.cell(row=2, column=col).value
        if cell_val is not None and str(cell_val).strip() == target_value:
            val_col = col
            break
    return date_col, val_col


def find_fallback_col_idx(ws, date_col, fallback_name):
    if not fallback_name or not date_col:
        return None
    target = str(fallback_name).strip()
    for col in range(max(1, date_col - 5), date_col + 50):
        cell_val = ws.cell(row=2, column=col).value
        if cell_val is not None and str(cell_val).strip() == target:
            return col
    return None


def read_series(ws, date_col, val_col, fallback_col=None):
    primary = {}
    fallback = {}
    null_streak = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_date = row[date_col - 1]
        if raw_date is None:
            null_streak += 1
            if null_streak >= 5:
                break
            continue
        null_streak = 0

        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            continue

        raw_val = row[val_col - 1]
        if raw_val is not None:
            try:
                primary[d] = float(raw_val)
            except (TypeError, ValueError):
                pass

        if fallback_col is not None:
            raw_fb = row[fallback_col - 1]
            if raw_fb is not None:
                try:
                    fallback[d] = float(raw_fb)
                except (TypeError, ValueError):
                    pass

    for d, v in fallback.items():
        primary.setdefault(d, v)
    return primary


def nearest_on_or_before(series, target):
    candidates = [d for d in series if d <= target]
    if not candidates: return None, None
    best = max(candidates)
    return best, series[best]

def calc_change(t0_val, ref_val, index_type):
    if t0_val is None or ref_val is None:
        return None, None
    change = t0_val - ref_val
    if index_type in ("rate", "spread"):
        return round(change, 6), round(change * 100, 2)
    pct = round(change / ref_val * 100, 3) if ref_val else None
    return round(change, 6), pct


def build_pending(cfg, errors):
    return {
        "label": cfg["label"],
        "section": cfg["section"],
        "type": cfg["type"],
        "holiday": False,
        "pending": True,
        "error": "; ".join(errors) if errors else None,
        "T0": {"date": None, "value": None},
        "1D": {"date": None, "value": None, "change": None},
        "1M": {"date": None, "value": None, "change": None},
        "YTM": {"date": None, "value": None, "change": None},
    }


def generate_data(excel_path, output_path=None,
                  override_date=None, d1_override=None,
                  ytm_override=None, generated_at_override=None):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"엑셀 없음: {excel_path}")

    t0_star = parse_iso_date(generated_at_override or date.today().isoformat(), "T0*")
    t0_date = parse_iso_date(override_date or (date.today() - timedelta(days=1)).isoformat(), "T0")
    t1_date = parse_iso_date(d1_override or (t0_date - timedelta(days=1)).isoformat(), "T-1")
    ytm_date = parse_iso_date(ytm_override or f"{t0_date.year}-01-02", "연초")
    one_m_date = t0_date - relativedelta(months=1)

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    results = []

    for cfg in INDEX_CONFIG:
        if cfg["sheet"] == "__pending__":
            results.append(build_pending(cfg, ["소스 미연결"]))
            continue

        if cfg["sheet"] not in wb.sheetnames:
            results.append(build_pending(cfg, [f"시트 없음: {cfg['sheet']}"]))
            continue

        ws = wb[cfg["sheet"]]
        date_col = val_col = None
        header_used = None
        for cand in cfg.get("header_candidates", [cfg["header"]]):
            date_col, val_col = find_columns(ws, cand, cfg["value_col"])
            if date_col and val_col:
                header_used = cand
                break

        if not (date_col and val_col):
            results.append(build_pending(cfg, [f"헤더/컬럼 탐지 실패: {cfg['header']} / {cfg['value_col']}"]))
            continue

        fallback_col = find_fallback_col_idx(ws, date_col, cfg.get("fallback_col"))
        series = read_series(ws, date_col, val_col, fallback_col=fallback_col)

        t0_val  = series.get(t0_date)
        t1_val  = series.get(t1_date)
        m1_date_actual, m1_val = nearest_on_or_before(series, one_m_date)
        # 지표별 연초 기준일 (개별 설정 없으면 공통 ytm_date 사용)
        cfg_ytm = cfg.get("ytm_date")
        effective_ytm = date.fromisoformat(cfg_ytm) if cfg_ytm else ytm_date
        ytm_val = series.get(effective_ytm)

        missing = []
        if t0_val is None:
            missing.append(f"T0 {t0_date} 데이터 없음")
        if t1_val is None:
            missing.append(f"T-1 {t1_date} 데이터 없음")
        if ytm_val is None:
            missing.append(f"연초 {effective_ytm} 데이터 없음")

        if missing:
            results.append(build_pending(cfg, missing))
            continue

        _, d1_change = calc_change(t0_val, t1_val, cfg["type"])
        _, m1_change = calc_change(t0_val, m1_val, cfg["type"])
        _, ytm_change = calc_change(t0_val, ytm_val, cfg["type"])

        results.append({
            "label": cfg["label"],
            "section": cfg["section"],
            "type": cfg["type"],
            "holiday": False,
            "pending": False,
            "error": None,
            "source_header": header_used,
            "T0": {"date": str(t0_date), "value": t0_val},
            "1D": {"date": str(t1_date), "value": t1_val, "change": d1_change},
            "1M": {"date": str(m1_date_actual) if m1_date_actual else None, "value": m1_val, "change": m1_change},
            "YTM": {"date": str(effective_ytm), "value": ytm_val, "change": ytm_change},
        })

    wb.close()

    chart_cutoff = t0_date - timedelta(days=366)
    wb2 = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    chart_series = []
    for ccfg in CHART_CONFIG:
        sd = {"label": ccfg["label"], "color": ccfg["color"], "dash": ccfg["dash"], "dates": [], "values": []}
        try:
            if ccfg["sheet"] in wb2.sheetnames:
                ws2 = wb2[ccfg["sheet"]]
                dc = vc = None
                for cand in ccfg.get("header_candidates", [ccfg["header"]]):
                    dc, vc = find_columns(ws2, cand, ccfg["value_col"])
                    if dc and vc:
                        break
                fbc = find_fallback_col_idx(ws2, dc, ccfg.get("fallback_col")) if dc else None
                if dc and vc:
                    raw = read_series(ws2, dc, vc, fallback_col=fbc)
                    pairs = sorted((d, v) for d, v in raw.items() if chart_cutoff <= d <= t0_date)
                    sd["dates"] = [str(d) for d, _ in pairs]
                    sd["values"] = [v for _, v in pairs]
        except Exception:
            pass
        chart_series.append(sd)
    wb2.close()

    output = {
        "generated_at": str(t0_star),
        "base_date": str(t0_date),
        "t_minus1": str(t1_date),
        "ytm_start": str(ytm_date),
        "today": str(date.today()),
        "indices": results,
        "chart_series": chart_series,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    return output


if __name__ == "__main__":
    import sys
    excel = sys.argv[1] if len(sys.argv) > 1 else "데이터.xlsx"
    generate_data(excel)
